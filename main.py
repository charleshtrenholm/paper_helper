#!/usr/bin/env python3

import os
import re
import shutil
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, colors
from openpyxl.utils import get_column_letter
from datetime import datetime
import glob
import subprocess

DOCUMENT_TYPES = [
    'Fieldprint',
    'TrueScreen',
    'MSA',
    'DL',
    'PP',
    'BC',
    'Mid Eval',
    'Final Eval',
    'Embark'
]

print('Moving labelled documents to their appropriate places...')

# create directories for document types if they don't already exist
base_dir = os.getcwd()
sorted_dir = os.path.join(base_dir, 'sorted_files')
for doctype in DOCUMENT_TYPES:
    if not os.path.exists(os.path.join(sorted_dir, doctype)):
        os.mkdir(os.path.join(sorted_dir, doctype))

moved = 0
not_moved = 0

# get user input list of interns
with open('interns.txt') as file:
    intern_names = file.readlines()
    intern_names = [line.rstrip() for line in intern_names]

fmt = r"[a-zA-Z]\.[a-zA-Z]+\s[a-zA-Z\s]+\.(?:jpg|png|pdf|jpeg|webp)"
for filename in os.listdir('new_files'):
    # check if it correctly matches file format
    if not re.fullmatch(fmt, filename):
        print('Incorrect file formatting for "%s", please rename appropriately' % filename)
        not_moved += 1
        continue

    name, file_type = filename.split(" ", 1)
    destination = file_type.split('.')[0]
    first_initial, last_name = name.split('.')

    # Check if intern name is on the list (or spelled wrong)
    if not any(x.startswith(first_initial) and x.endswith(last_name) for x in intern_names):
        print('No intern exitsts with the name "%s", skipping "%s"' % (name, filename))
        not_moved += 1
        continue

    # Check if theres a weird random document type that shouldn't exist
    if not any(destination.lower() == d.lower() for d in DOCUMENT_TYPES):
        print('Found file "%s" with improper document type "%s", what the HELL are you doing' % (filename, destination))
        not_moved += 1
        continue

    # actually move the bad boi
    shutil.move(
        os.path.join(base_dir, 'new_files', filename),
        os.path.join(sorted_dir, destination, filename)
    )

    moved += 1


print('Successfully moved %s files into their proper place - %s unsorted files remain' % (moved, not_moved))

def fmt_intern_name(intern_name):
    firstname, lastname = intern_name.split(' ')
    return firstname[0] + '.' + lastname

def as_text(value):
    if value is None:
        return ''
    return str(value)

# now time to make the excel sheet
wb = Workbook()
ws = wb.active
now = datetime.now().strftime("%d_%m_%Y_%H:%M:%S")
sheet_name = 'intern_papers_' + now + '.xlsx'

HEADER_FONT = Font(name="Calibri", size=14, bold=True, extend=True)
FILE_EXISTS_FILL = PatternFill('solid', fgColor="CCFFCC")
FILE_NOT_FOUND_FILL = PatternFill('solid', fgColor="FFCCCC")
# FFCCCC = red
# ccffcc = green
# create sheet headers
ws.append(['Intern Name'] + DOCUMENT_TYPES)

for cell in ws['1:1']:
    cell.font = HEADER_FONT

for intern_name in intern_names:
    name = fmt_intern_name(intern_name)
    turned_in = []
    for doctype in DOCUMENT_TYPES:
        filename = f'{name} {doctype}'
    
        if any(os.path.splitext(x)[0].lower() == filename.lower() for x in os.listdir(os.path.join(sorted_dir, doctype))):
            turned_in.append('yes')
        else:
            turned_in.append('no')

    ws.append([intern_name] + turned_in)

# set width of columns
for column_cells in ws.columns:
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = 20

# make the cells pretty colors
for row in ws.iter_rows(min_row=2, max_row=len(intern_names) + 1, min_col=2, max_col=len(DOCUMENT_TYPES) + 1):
    for cell in row:
        if cell.value == 'no':
            cell.fill = FILE_NOT_FOUND_FILL
        elif cell.value == 'yes':
            cell.fill = FILE_EXISTS_FILL
        else:
            cell.fill = PatternFill('solid', fgColor="FFFFCC")

# save workbook and open it
wb.save(os.path.join(base_dir, 'results', sheet_name))
subprocess.run('open results/%s' % sheet_name, shell=True)



